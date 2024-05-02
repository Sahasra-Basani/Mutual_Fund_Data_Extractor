import os.path
import openpyxl
import process_date
from database.database import connect_to_db
import json

with open("../config.json") as f:
    config = json.load(f)

# Current date and time
current_date = process_date.current_date

# Get the previous month
prev_mon_name_MM = process_date.prev_mon_name_MM

# Get the current year four digits
current_year_yyyy = process_date.current_year_yyyy

# Get the days in the month
days_in_month = process_date.days_in_month


def insert_in_stock_table():
    category = ""
    fund_id = 0

    try:
        # Input File path
        # file_path = "F:/Python Projects/Funds details/Nippon/NIMF-MONTHLY-PORTFOLIO-REPORT-March-24.xls"
        file_name = config["other_settings"]["excel_path_sbi"]
        file_path = file_name+f"{days_in_month}-{prev_mon_name_MM.lower()}-{current_year_yyyy}.xlsx"
        print("file_path", file_path)

        if not os.path.exists(file_path):
            print("File not found!")
            exit()
        else:
            mon_yr = prev_mon_name_MM+'_'+str(current_year_yyyy)

            # Sheets to be read from excel
            sheet_names = ['SSCF', 'SMIDCAP', 'SLMF', 'SLTEF']

            for sheet_name in sheet_names:

                # Open the Excel file
                workbook = openpyxl.load_workbook(file_path)

                # Select a worksheet
                worksheet = workbook[sheet_name]

                # Insert data in the fund_details table
                insert_in_fund_table(worksheet)

                # Define the list of column indices you want to read
                columns_to_read = [2, 3, 4, 7]  # Columns 3, 4, 5, and 8 (0-indexed)

                # Define the starting row 10
                start_row = 9  # Row 10 (0-indexed)

                # Iterate over rows starting from the 10th row
                for row in worksheet.iter_rows(min_row=10, values_only=True):
                    # Extract data from the row
                    column1_value = row[2]  # Assuming column 3 is the first column
                    column2_value = row[3]  # Assuming column 4 is the second column
                    column3_value = row[4]  # Assuming column 5 is the third column
                    column6_value = row[7]  # Assuming column 8 is the sixth column

                    # Check if the value in column 4 is not null
                    if row[3] is not None and row[3].startswith('IN'):

                        data = worksheet.cell(row=3, column=4).value

                        if "Smallcap" in data:
                            category = "Small Cap"
                        elif "Magnum Midcap" in data:
                            category = "Mid Cap"
                        elif "Large" in data:
                            category = "Large Cap"
                        elif "Long Term Equity" in data:
                            category = "Tax Saver"

                        # Get fund_id from fund_details table
                        fund_id = get_fund_id(data)

                        try:
                            # Insert data into the database in stock_details table
                            cursor.execute('''INSERT INTO mutual_fund_data.stock_details(fund_id, created_date, mon_yr,
                                                    category, stock_name, isin_no , industry, holding_share)
                                                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s)''',
                                           (fund_id,
                                            current_date,
                                            mon_yr,
                                            category,
                                            column1_value,
                                            column2_value,
                                            column3_value,
                                            column6_value))

                        except Exception as e:
                            print("Issue occurred inserting dta into database in stock_details table!!!", str(e))

                print("Data inserted successfully in the stock_details table for {} category!!!\n".format(category))

    except Exception as e:
        print("Issue while working on reading excel", str(e))


def insert_in_fund_table(worksheet):
    try:
        # fund_house and created_date values
        fund_house = "SBI Mutual Fund"

        # Insert the data into the database in fund_details table
        cursor.execute('''INSERT INTO mutual_fund_data.fund_details(created_date, fund_name, fund_house)
        VALUES (%s, %s, %s)''',
                       (current_date,
                        worksheet.cell(row=3, column=4).value,
                        fund_house))
        print("Data inserted successfully in the fund_details table!!!")

    except Exception as e:
        print("Issue occurred inserting dta into database in fund_details table!!!", str(e))


def get_fund_id(category):
    try:
        # Get fund_id from fund_details table
        query = ("SELECT id FROM mutual_fund_data.fund_details WHERE fund_house = 'SBI Mutual Fund' "
                 "and lower(fund_name) LIKE %s LIMIT 1")
        cursor.execute(query, ('%' + category.lower() + '%',))
        res = cursor.fetchone()
        return res

    except Exception as e:
        print("Issue occurred while getting the fund_id from the fund_details table!!!", str(e))


# Database connection details
host = config["database"]["host"]
port = config["database"]["port"]
dbname = config["database"]["database_name"]
user = config["database"]["username"]
password = config["database"]["password"]

try:
    # Connect to the PostgreSQL database
    connection = connect_to_db(dbname, user, password, host, port)

    if connection:
        cursor = connection.cursor()
        print("**************Connection Successful*******************")

        insert_in_stock_table()

except Exception as e:
    print("Issue occurred while connecting to the database", str(e))

finally:
    # Commit changes and close the connection
    connection.commit()
    connection.close()



